from openpyxl import Workbook

first_ex = Workbook()
file_path = '../excel/grade.xlsx'

first_ex_at = first_ex.active
first_ex_at.title = '学生成绩表'

first_ex_at['A1'] = '学号'
first_ex_at['B1'] = '姓名'
first_ex_at['C1'] = '语文'
first_ex_at['D1'] = '数学'
first_ex_at['E1'] = '英语'

student = [
    ("202401", "张明", 88, 95, 92),
    ("202402", "李娜", 92, 89, 96),
    ("202403", "王浩", 79, 98, 85),
    ("202404", "刘阳", 95, 93, 90),
    ("202405", "陈雨", 85, 82, 88),
    ("202406", "赵宇", 90, 97, 94),
    ("202407", "孙萌", 82, 78, 83),
    ("202408", "周凯", 96, 91, 98),
    ("202409", "吴婷", 87, 86, 91),
    ("202410", "郑杰", 75, 92, 79),
    ("202411", "钱悦", 93, 88, 95),
    ("202412", "冯磊", 81, 85, 80),
    ("202413", "高思", 97, 99, 96),
    ("202414", "韩雪", 89, 80, 87),
    ("202415", "朱晨", 78, 90, 82),
    ("202416", "马丽", 91, 94, 89),
    ("202417", "胡强", 84, 87, 93),
    ("202418", "林晓", 94, 83, 91),
    ("202419", "郭涛", 80, 96, 86),
    ("202420", "何雅", 86, 89, 97)
]

for index, data in enumerate(student, start=2):
    first_ex_at.cell(row = index, column=1, value = data[0])
    first_ex_at.cell(row = index, column=2, value = data[1])
    first_ex_at.cell(row = index, column=3, value = data[2])
    first_ex_at.cell(row = index, column = 4, value = data[3])
    first_ex_at.cell(row = index, column = 5, value = data[4])

first_ex.save(file_path)

print('ok')